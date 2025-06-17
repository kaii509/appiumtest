from appium.webdriver.common.appiumby import AppiumBy
from openpyxl import load_workbook
import time

app_package = "com.chikkanna.tgcloginapp"
path = "/Users/khus1en/Documents/Internship Xac/PyAppiumTest/tests/credentials.xlsx"


wb = load_workbook(path)
sheet = wb.active

sheet.cell(row=1, column=3).value = "Result"

def is_login_successful(driver):
    try:
        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/successMsg")')
        return True
    except:
        return False
    

def test_login(driver):
    # Iteration assuming that the first col contains the username and, the second col contains the password
    for row in range(2, sheet.max_row + 1):


        username = sheet.cell(row=row, column=1).value
        password = sheet.cell(row=row, column=2).value


        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/usernameField")').clear()
        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/usernameField")').send_keys(username)

        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/passwordField")').clear()
        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/passwordField")').send_keys(password)

        driver.find_element(AppiumBy.ANDROID_UIAUTOMATOR,'new UiSelector().resourceId("com.chikkanna.tgcloginapp:id/loginBtn")').click()

        time.sleep(2)

        if is_login_successful(driver): result = "passed"
        else: result = "failed"
        sheet.cell(row=row, column=3).value = result
  


        driver.terminate_app(app_package)
        driver.activate_app(app_package)
    wb.save(path)
    wb.close() 

